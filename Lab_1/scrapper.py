import json
import openpyxl
import csv
from bs4 import BeautifulSoup


def create_json(data: list) -> None:
    with open('outputs/output.json', 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)


def create_xlsx(data: list) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Inflation Data'

    sheet['A1'] = 'date'
    sheet['B1'] = 'actual_inflation'

    for row_index, entry in enumerate(data, start=2):
        sheet.cell(row=row_index, column=1, value=entry['date'])
        sheet.cell(row=row_index, column=2, value=entry['actual_inflation'])

    workbook.save('outputs/output.xlsx')


def create_csv(data: list) -> None:
    with open('outputs/output.csv', 'w', newline='', encoding='utf-8') as csv_file:
        csv_writer = csv.writer(csv_file)
        csv_writer.writerow(['date', 'actual_inflation'])

        for entry in data:
            csv_writer.writerow([entry['date'], entry['actual_inflation']])


def scrap_file(create_file=True):
    with open('output.html', 'r', encoding='utf-8') as html_file:
        html_content = html_file.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.find('table', {'id': 'eventHistoryTable733'})
    data = []

    for row in table.find_all('tr', {'event_attr_id': '733'}):
        cells = row.find_all('td')
        if len(cells) == 6:
            date_str = cells[0].text.strip()
            date_str = date_str[0:10]
            actual_inflation = cells[2].text.strip()
            data.append({'date': date_str, 'actual_inflation': actual_inflation})
    if create_file:
        create_json(data=data)
        create_csv(data=data)
        create_xlsx(data=data)
    else:
        return data


scrap_file()
